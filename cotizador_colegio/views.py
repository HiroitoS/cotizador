# ✅ ARCHIVO: cotizador_colegio/views.py
# -----------------------------------------------------------------------------
# NOTA IMPORTANTE (qué se corrigió respecto al 29 de octubre):
# 1) GuardarCotizacionView:
#    - Ya NO llama internamente a CalcularDetalleView con request falso (eso causaba 500).
#    - Ahora calcula servidor-side:
#       * Comisión en CONSIGNA = MONTO FIJO (no %)
#       * ROI nunca negativo
#    - Si no llega tipo_venta en el body raíz, toma del primer detalle.
# 2) CotizacionPanelSerializer usado para el Panel. Normaliza 'PV' -> 'PUNTO_DE_VENTA'.
# 3) DetalleCotizacionRetrieveView:
#    - Devuelve estructura para el ModalAdopcion (libro/área/grado + es_plan_lector).
# 4) Exportaciones Excel:
#    - 'safe()' evita negativos en celdas.
# 5) CalcularDetalleView:
#    - Mismo criterio: comisión = monto fijo; ROI >= 0; se mantiene para uso del front al calcular en vivo.
# -----------------------------------------------------------------------------

from io import BytesIO
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.db.models.functions import Lower
from django.conf import settings

from django.utils import timezone
from django.views import View

from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import generics, filters, status
from rest_framework.generics import ListAPIView
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.decorators import api_view

from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import (
    Cotizacion, Libro, Adopcion, Pedido,
    AsesorComercial, InstitucionEducativa,
    DetalleCotizacion, DetalleAdopcion, 
)
from .serializers import (
    
    CotizacionDetalleSerializer, CotizacionPanelSerializer, LibroSerializer, PedidoSerializer

    , AsesorSerializer, InstitucionSerializer, AdopcionPanelSerializer
)
from .services_pdf import generar_pdf_cotizacion, generar_pdf_adopcion


# -------------------------
# Utilidades de redondeo/seguridad
# -------------------------
TWO = Decimal("0.01")
ZERO = Decimal("0.00")
HUNDRED = Decimal("100")

def nz_decimal(value):
    """Convierte a Decimal con 2 decimales, evitando None/NaN/negativos."""
    try:
        d = Decimal(str(value))
        if d < ZERO:
            d = ZERO
        return d.quantize(TWO, rounding=ROUND_HALF_UP)
    except Exception:
        return ZERO

def clamp_non_negative(d: Decimal) -> Decimal:
    return d if d >= ZERO else ZERO


# -------------------------
# 1) LIBROS con filtros/búsqueda
# -------------------------
class ListarLibrosView(generics.ListAPIView):
    queryset = Libro.objects.all().order_by("empresa", "nivel", "grado", "area")
    serializer_class = LibroSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ["nivel", "area", "grado", "empresa"]
    search_fields = ["descripcion_completa", "serie", "empresa"]


# -------------------------
# 2) GUARDAR COTIZACIÓN
# -------------------------
class GuardarCotizacionView(APIView):
    """
    Espera payload:
    {
      "institucion": {"id": ...},
      "asesor": {"id": ...},
      "tipo_venta": "PV" | "FERIA" | "CONSIGNA",   # <- si no llega, se toma del 1er detalle
      "detalles": [
        {
          "libro_id": 1,
          "tipo_venta": "PV"|"FERIA"|"CONSIGNA",   # opcional si ya viene en raíz
          "precio_be": 200,
          "desc_proveedor": 10,                    # %
          "precio_ie": 160, "precio_ppff": 180,    # PV/FERIA
          "desc_consigna": 15, "comision": 20      # CONSIGNA (comisión = MONTO FIJO)
        }, ...
      ]
    }
    """
    def post(self, request):
        data = request.data
        institucion = data.get("institucion") or {}
        asesor = data.get("asesor") or {}
        detalles = data.get("detalles") or []
        tipo_venta_root = (data.get("tipo_venta") or "").upper().strip()

        if not institucion.get("id"):
            return Response({"detail": "Falta institución."}, status=400)
        if not asesor.get("id"):
            return Response({"detail": "Falta asesor."}, status=400)
        if not isinstance(detalles, list) or not detalles:
            return Response({"detail": "Faltan detalles."}, status=400)

        # Si no viene en raíz, toma del primer detalle
        if not tipo_venta_root:
            tipo_venta_root = (detalles[0].get("tipo_venta") or "").upper().strip()

        cot = Cotizacion.objects.create(
            institucion_id=institucion["id"],
            asesor_id=asesor["id"],
            
        )

        # Cálculo servidor-side (comisión fija en CONSIGNA; ROI >= 0)
        for det in detalles:
            tv = (det.get("tipo_venta") or tipo_venta_root or "").upper().strip()
            if tv not in ("PV", "FERIA", "CONSIGNA"):
                tv = "PV"

            precio_be = nz_decimal(det.get("precio_be"))
            # regla: precio_be >= pvp_2026
            libro = Libro.objects.get(id=det["libro_id"])
            if precio_be < libro.pvp_2026_con_igv:
                precio_be = libro.pvp_2026_con_igv

            desc_proveedor_pct = nz_decimal(det.get("desc_proveedor"))
            precio_ie = nz_decimal(det.get("precio_ie")) if det.get("precio_ie") is not None else None
            precio_ppff = nz_decimal(det.get("precio_ppff")) if det.get("precio_ppff") is not None else None
            desc_consigna_pct = nz_decimal(det.get("desc_consigna")) if det.get("desc_consigna") is not None else None
            comision_monto = nz_decimal(det.get("comision")) if det.get("comision") is not None else None

            # precio proveedor
            precio_proveedor = (precio_be * (Decimal("1") - (desc_proveedor_pct / HUNDRED))).quantize(TWO, rounding=ROUND_HALF_UP)
            if precio_proveedor < ZERO:
                precio_proveedor = ZERO

            # Inicializa campos
            precio_consigna = None
            precio_coordinado = None
            utilidad_ie = ZERO
            roi_ie = None
            roi_consigna = None

            if tv in ("PV", "FERIA"):
                if precio_ie is not None and precio_ppff is not None:
                    utilidad_ie = (precio_ppff - precio_ie).quantize(TWO, rounding=ROUND_HALF_UP)
                    roi_ie_raw = (precio_ie - precio_proveedor).quantize(TWO, rounding=ROUND_HALF_UP)
                    roi_ie = clamp_non_negative(roi_ie_raw)
                # limpiar consignación
                desc_consigna_val = None
                comision_val = None
            else:
                # CONSIGNA: desc_consigna = %, comision = MONTO FIJO
                desc_c = (desc_consigna_pct or ZERO) / HUNDRED
                precio_consigna = (precio_be * (Decimal("1") - desc_c)).quantize(TWO, rounding=ROUND_HALF_UP)
                if precio_consigna < ZERO:
                    precio_consigna = ZERO

                com_monto = (comision_monto or ZERO)
                if com_monto < ZERO:
                    com_monto = ZERO

                precio_coordinado = (precio_consigna - com_monto).quantize(TWO, rounding=ROUND_HALF_UP)
                if precio_coordinado < ZERO:
                    precio_coordinado = ZERO

                utilidad_ie = (precio_be - precio_consigna).quantize(TWO, rounding=ROUND_HALF_UP)
                roi_raw = (precio_coordinado - precio_proveedor).quantize(TWO, rounding=ROUND_HALF_UP)
                roi_consigna = clamp_non_negative(roi_raw)

            DetalleCotizacion.objects.create(
                cotizacion=cot,
                libro_id=det["libro_id"],
                tipo_venta=tv,
                # guardamos lo que ingresa usuario (no negativos)
                precio_be=precio_be,
                desc_proveedor=desc_proveedor_pct,
                precio_ie=precio_ie,
                precio_ppff=precio_ppff,
                desc_consigna=desc_consigna_pct,
                comision=comision_monto,
                # derivados
                precio_proveedor=precio_proveedor,
                precio_consigna=precio_consigna,
                precio_coordinado=precio_coordinado,
                utilidad_ie=utilidad_ie,
                roi_ie=roi_ie,
                roi_consigna=roi_consigna,
            )

        return Response({"numero": cot.numero_cotizacion}, status=201)


# -------------------------
# 3) LISTAR COTIZACIONES (panel)
# -------------------------
class ListarCotizacionesView(APIView):
    def get(self, request):
        qs = Cotizacion.objects.select_related("institucion", "asesor").prefetch_related("detalles")
        data = CotizacionPanelSerializer(qs.order_by("-id"), many=True).data
        return Response(data, status=status.HTTP_200_OK)


# -------------------------
# 4) CAMBIAR ESTADO (PENDIENTE -> APROBADA/RECHAZADA)
# -------------------------
class CambiarEstadoCotizacionView(APIView):
    """
    PATCH /api/cotizaciones/estado/<id>/
    Body: { "estado": "APROBADA" | "RECHAZADA", "motivo": "..." }
    """
    def patch(self, request, pk):
        cot = get_object_or_404(Cotizacion, pk=pk)
        estado = request.data.get("estado")
        motivo = (request.data.get("motivo") or "").strip()

        if estado not in ["APROBADA", "RECHAZADA"]:
            return Response({"detail": "Estado no válido."}, status=400)

        cot.estado = estado
        if hasattr(cot, "motivo_rechazo"):
            cot.motivo_rechazo = motivo if estado == "RECHAZADA" else None

        fields = ["estado"]
        if hasattr(cot, "motivo_rechazo"):
            fields.append("motivo_rechazo")
        cot.save(update_fields=fields)

        return Response({
            "message": f"Estado de la cotización {cot.numero_cotizacion or cot.id} actualizado correctamente.",
            "estado": cot.estado,
            "motivo_rechazo": getattr(cot, "motivo_rechazo", None),
        }, status=200)


# -------------------------
# 5) CALCULAR DETALLE (para cálculo en vivo en el front)
# -------------------------
class CalcularDetalleView(APIView):
    """
    Calcula precios/utilidades/ROI según tipo_venta.
    CONSIGNA: comisión = MONTO FIJO.
    ROI nunca negativo.
    """
    def post(self, request, *args, **kwargs):
        data = request.data or {}
        libro_id = data.get("libro") or data.get("id")
        tipo_venta = (data.get("tipo_venta") or "").upper().strip()

        if not libro_id or not tipo_venta:
            return Response({"detail": "Campos requeridos: libro, tipo_venta."}, status=400)

        libro = get_object_or_404(Libro, pk=libro_id)

        precio_be = nz_decimal(data.get("precio_be", libro.pvp_2026_con_igv or 0))
        desc_proveedor_pct = nz_decimal(data.get("desc_proveedor", 0))
        precio_proveedor = (precio_be * (Decimal("1") - (desc_proveedor_pct / HUNDRED))).quantize(TWO, rounding=ROUND_HALF_UP)
        if precio_proveedor < ZERO:
            precio_proveedor = ZERO

        resp = {
            "id": libro.id,
            "precio_be": precio_be,
            "desc_proveedor": desc_proveedor_pct,
            "precio_proveedor": precio_proveedor,
        }

        if tipo_venta in ("PV", "FERIA"):
            precio_ie = nz_decimal(data.get("precio_ie", 0))
            precio_ppff = nz_decimal(data.get("precio_ppff", 0))

            utilidad_ie = (precio_ppff - precio_ie).quantize(TWO, rounding=ROUND_HALF_UP)
            roi_raw = (precio_ie - precio_proveedor).quantize(TWO, rounding=ROUND_HALF_UP)

            resp.update({
                "tipo_venta": tipo_venta,
                "precio_ie": precio_ie,
                "precio_ppff": precio_ppff,
                "utilidad_ie": utilidad_ie,
                "roi_ie": clamp_non_negative(roi_raw),
                "precio_consigna": None,
                "precio_coordinado": None,
                "roi_consigna": None,
                "desc_consigna": None,
                "comision": None,
            })
            return Response(resp, status=200)

        elif tipo_venta == "CONSIGNA":
            desc_consigna_pct = nz_decimal(data.get("desc_consigna", 0))
            comision_monto = nz_decimal(data.get("comision", 0))  # MONTO

            if desc_consigna_pct > Decimal("50"):
                return Response({"detail": "El descuento de consignación no puede superar 50%."}, status=400)

            desc_c = (desc_consigna_pct / HUNDRED)
            precio_consigna = (precio_be * (Decimal("1") - desc_c)).quantize(TWO, rounding=ROUND_HALF_UP)
            if precio_consigna < ZERO:
                precio_consigna = ZERO

            precio_coordinado = (precio_consigna - comision_monto).quantize(TWO, rounding=ROUND_HALF_UP)
            if precio_coordinado < ZERO:
                precio_coordinado = ZERO

            utilidad_ie = (precio_be - precio_consigna).quantize(TWO, rounding=ROUND_HALF_UP)
            roi_raw = (precio_coordinado - precio_proveedor).quantize(TWO, rounding=ROUND_HALF_UP)

            resp.update({
                "tipo_venta": tipo_venta,
                "desc_consigna": desc_consigna_pct,
                "comision": comision_monto,
                "precio_consigna": precio_consigna,
                "precio_coordinado": precio_coordinado,
                "utilidad_ie": utilidad_ie,
                "roi_consigna": clamp_non_negative(roi_raw),
                "precio_ie": None,
                "precio_ppff": None,
                "roi_ie": None,
            })
            return Response(resp, status=200)

        return Response({"detail": "tipo_venta inválido."}, status=400)


# -------------------------
# 6) DETALLE de cotización (para Modal Adopción)
# -------------------------
class DetalleCotizacionRetrieveView(APIView):
    def get(self, request, pk):
        cot = get_object_or_404(Cotizacion, id=pk)
        # Usamos serializer específico para el modal
        return Response(CotizacionDetalleSerializer(cot).data, status=200)


# -------------------------
# 7) Crear Adopción
# -------------------------
class CrearAdopcionView(APIView):
    """
    POST /api/adopciones/crear/
    payload: { cotizacion_id, items: [{detalle_id, cantidad, mes_lectura?}] }
    """
    def post(self, request):
        cotizacion_id = request.data.get("cotizacion_id")
        items = request.data.get("items") or []
        if not cotizacion_id or not items:
            return Response({"detail": "cotizacion_id e items son requeridos."}, status=400)

        cot = get_object_or_404(Cotizacion, pk=cotizacion_id)

        adop, _ = Adopcion.objects.get_or_create(cotizacion=cot)
        adop.detalles.all().delete()

        for it in items:
            det_id = it.get("detalle_id")
            cant = it.get("cantidad")
            mes = it.get("mes_lectura")

            if not det_id or not cant:
                return Response({"detail": "Cada item requiere detalle_id y cantidad > 0."}, status=400)

            det_cot = get_object_or_404(
                DetalleCotizacion.objects.select_related("libro"),
                pk=det_id, cotizacion=cot
            )
            DetalleAdopcion.objects.create(
                adopcion=adop,
                libro=det_cot.libro,
                cantidad_adoptada=int(cant),
                mes_lectura=mes
            )

        if cot.estado != "ADOPTADA":
            cot.estado = "ADOPTADA"
            cot.save(update_fields=["estado"])

        return Response({"adopcion_id": adop.id}, status=201)


class ExportarAdopcionPDFView(APIView):
    """Devuelve el PDF de una ficha de adopción específica (por id de adopción)."""
    def get(self, request, adopcion_id):
        try:
            adopcion = Adopcion.objects.get(id=adopcion_id)
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="Adopcion_{adopcion_id}.pdf"'
            generar_pdf_adopcion(adopcion, response)
            return response
        except Adopcion.DoesNotExist:
            return Response({"detail": "Ficha de adopción no encontrada."}, status=404)
        except Exception as e:
            print("Error al generar PDF de adopción:", e)
            return Response({"detail": "Error interno al generar el PDF."}, status=500)

# ---------------------------
# LISTAR ADOPCIONES – NECESARIO PARA PANEL ADOPCIONES
# ---------------------------
class ListarAdopcionesView(APIView):
    def get(self, request):
        qs = (
            Adopcion.objects
            .select_related("cotizacion", "cotizacion__institucion", "cotizacion__asesor")
            .prefetch_related("cotizacion__detalles")
            .order_by("-id")
        )

        data = AdopcionPanelSerializer(qs, many=True).data
        return Response(data, status=200)



# -------------------------
# 8) PEDIDOS (listado)
# -------------------------
class ListarPedidosView(generics.ListAPIView):
    queryset = Pedido.objects.select_related("adopcion", "adopcion__cotizacion").order_by("-fecha_pedido")
    serializer_class = PedidoSerializer


# -------------------------
# 9) Filtros de catálogos
# -------------------------
@api_view(["GET"])
def filtros_libros(request):
    data = {
        "empresas": list(
            Libro.objects.exclude(empresa__isnull=True)
            .exclude(empresa__exact="")
            .values_list("empresa", flat=True)
            .distinct()
            .order_by(Lower("empresa"))
        ),
        "niveles": list(
            Libro.objects.exclude(nivel__isnull=True)
            .exclude(nivel__exact="")
            .values_list("nivel", flat=True)
            .distinct()
            .order_by(Lower("nivel"))
        ),
        "areas": list(
            Libro.objects.exclude(area__isnull=True)
            .exclude(area__exact="")
            .values_list("area", flat=True)
            .distinct()
            .order_by(Lower("area"))
        ),
        "grados": list(
            Libro.objects.exclude(grado__isnull=True)
            .exclude(grado__exact="")
            .values_list("grado", flat=True)
            .distinct()
            .order_by(Lower("grado"))
        ),
    }
    return Response(data)


# -------------------------
# 10) Asesores / Instituciones
# -------------------------
class ListarAsesoresView(ListAPIView):
    queryset = AsesorComercial.objects.filter(estado="ACTIVO").order_by("nombre")
    serializer_class = AsesorSerializer


class ListarColegiosView(ListAPIView):
    queryset = InstitucionEducativa.objects.all().order_by("nombre")
    serializer_class = InstitucionSerializer


# -------------------------
# 11) PDFs (cotización/adopción)
# -------------------------
class PDFCotizacionView(APIView):
    def get(self, request, pk):
        cot = get_object_or_404(Cotizacion, pk=pk)
        response = HttpResponse(content_type="application/pdf")
        response['Content-Disposition'] = f'attachment; filename="Cotizacion_{cot.numero_cotizacion}.pdf"'
        generar_pdf_cotizacion(cot, response)
        return response


class PDFAdopcionView(APIView):
    """Genera PDF de adopción por ID de cotización."""
    def get(self, request, pk):
        cot = get_object_or_404(Cotizacion, pk=pk)
        adop = get_object_or_404(Adopcion, cotizacion=cot)
        response = HttpResponse(content_type="application/pdf")
        response['Content-Disposition'] = f'attachment; filename="Adopcion_{cot.numero_cotizacion}.pdf"'
        generar_pdf_adopcion(adop, response)
        return response


# -------------------------
# 12) Exportaciones Excel (3 reportes)
# -------------------------
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
header_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
)

def estilo_encabezados(ws, fila=1):
    for cell in ws[fila]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

def ExportFilename(prefix):
    return f'{prefix}_{timezone.now().date()}.xlsx'

def safe(value):
    """Evita negativos y None en Excel."""
    if value is None:
        return ""
    if isinstance(value, (int, float, Decimal)) and value < 0:
        return 0
    return value

class ExportCotizacionesExcelView(View):
    def get(self, request, *args, **kwargs):
        detalles = DetalleCotizacion.objects.select_related("cotizacion", "libro").all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Cotizaciones"

        columnas = [
            "EMPRESA", "NIVEL", "GRADO", "ÁREA", "SERIE", "DESCRIPCIÓN COMPLETA",
            "TIPO DE INV", "SOPORTE", "PVP 2026 CON IGV", "DESC PROVEEDOR",
            "PRECIO PROVEEDOR", "TIPO DE VENTA", "PRECIO BE", "PRECIO IE",
            "PRECIO CONSIGNA", "PRECIO COORDINADO", "PRECIO PP.FF. (FERIA -PV)",
            "DESC CONSIGNA", "COMISIÓN", "UTILIDAD IE", "ROI X PROD. (PRECIO IE)",
            "ROI X PROD. (PRECIO CONSIGNA)", "ASESOR", "INSTITUCIÓN", "FECHA"
        ]
        ws.append(columnas)
        estilo_encabezados(ws)

        for d in detalles:
            libro = d.libro
            cot = d.cotizacion
            ws.append([
                safe(libro.empresa), safe(libro.nivel), safe(libro.grado), safe(libro.area), safe(libro.serie),
                safe(libro.descripcion_completa), safe(libro.tipo_inventario), safe(libro.soporte),
                safe(libro.pvp_2026_con_igv), safe(d.desc_proveedor), safe(d.precio_proveedor),
                safe(d.tipo_venta), safe(d.precio_be), safe(d.precio_ie), safe(d.precio_consigna),
                safe(d.precio_coordinado), safe(d.precio_ppff), safe(d.desc_consigna), safe(d.comision),
                safe(d.utilidad_ie), safe(d.roi_ie), safe(d.roi_consigna),
                safe(cot.asesor.nombre if cot.asesor else ""),
                safe(cot.institucion.nombre if cot.institucion else ""),
                safe(cot.fecha.strftime("%d/%m/%Y") if cot.fecha else ""),
            ])

        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{ExportFilename("Reporte_Cotizaciones")}"'
        wb.save(response)
        return response


class ExportAdopcionesExcelView(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        ws.title = "Adopciones"

        columnas = [
            "N° COTIZACIÓN", "INSTITUCIÓN", "ASESOR", "LIBRO", "EMPRESA",
            "NIVEL", "GRADO", "ÁREA", "CANTIDAD ADOPTADA", "MES DE LECTURA"
        ]
        ws.append(columnas)
        estilo_encabezados(ws)

        adopciones = Adopcion.objects.select_related(
            "cotizacion__institucion", "cotizacion__asesor"
        ).prefetch_related("detalles__libro").all()

        for adop in adopciones:
            for det in adop.detalles.all():
                ws.append([
                    safe(adop.cotizacion.numero_cotizacion),
                    safe(adop.cotizacion.institucion.nombre if adop.cotizacion.institucion else ""),
                    safe(adop.cotizacion.asesor.nombre if adop.cotizacion.asesor else ""),
                    safe(det.libro.descripcion_completa),
                    safe(det.libro.empresa), safe(det.libro.nivel), safe(det.libro.grado), safe(det.libro.area),
                    safe(det.cantidad_adoptada),
                    safe(det.mes_lectura or ""),
                ])

        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{ExportFilename("Reporte_Adopciones")}"'
        wb.save(response)
        return response


class ExportGeneralExcelView(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()

        # Hoja 1 - Cotizaciones
        ws1 = wb.active
        ws1.title = "Cotizaciones"

        columnas_cot = [
            "EMPRESA", "NIVEL", "GRADO", "ÁREA", "SERIE", "DESCRIPCIÓN COMPLETA",
            "TIPO DE INV", "SOPORTE", "PVP 2026 CON IGV", "DESC PROVEEDOR",
            "PRECIO PROVEEDOR", "TIPO DE VENTA", "PRECIO BE", "PRECIO IE",
            "PRECIO CONSIGNA", "PRECIO COORDINADO", "PRECIO PP.FF. (FERIA -PV)",
            "DESC CONSIGNA", "COMISIÓN", "UTILIDAD IE", "ROI X PROD. (PRECIO IE)",
            "ROI X PROD. (PRECIO CONSIGNA)", "ASESOR", "INSTITUCIÓN", "FECHA"
        ]
        ws1.append(columnas_cot)
        estilo_encabezados(ws1)

        detalles = DetalleCotizacion.objects.select_related("cotizacion", "libro").all()
        for d in detalles:
            libro = d.libro
            cot = d.cotizacion
            ws1.append([
                safe(libro.empresa), safe(libro.nivel), safe(libro.grado), safe(libro.area), safe(libro.serie),
                safe(libro.descripcion_completa), safe(libro.tipo_inventario), safe(libro.soporte),
                safe(libro.pvp_2026_con_igv), safe(d.desc_proveedor), safe(d.precio_proveedor),
                safe(d.tipo_venta), safe(d.precio_be), safe(d.precio_ie), safe(d.precio_consigna),
                safe(d.precio_coordinado), safe(d.precio_ppff), safe(d.desc_consigna), safe(d.comision),
                safe(d.utilidad_ie), safe(d.roi_ie), safe(d.roi_consigna),
                safe(cot.asesor.nombre if cot.asesor else ""),
                safe(cot.institucion.nombre if cot.institucion else ""),
                safe(cot.fecha.strftime("%d/%m/%Y") if cot.fecha else ""),
            ])

        # Hoja 2 - Adopciones
        ws2 = wb.create_sheet("Adopciones")
        columnas_adop = [
            "N° COTIZACIÓN", "INSTITUCIÓN", "ASESOR", "LIBRO", "EMPRESA",
            "NIVEL", "GRADO", "ÁREA", "CANTIDAD ADOPTADA", "MES DE LECTURA"
        ]
        ws2.append(columnas_adop)
        estilo_encabezados(ws2)

        adopciones = Adopcion.objects.select_related(
            "cotizacion__institucion", "cotizacion__asesor"
        ).prefetch_related("detalles__libro").all()

        for adop in adopciones:
            for det in adop.detalles.all():
                ws2.append([
                    safe(adop.cotizacion.numero_cotizacion),
                    safe(adop.cotizacion.institucion.nombre if adop.cotizacion.institucion else ""),
                    safe(adop.cotizacion.asesor.nombre if adop.cotizacion.asesor else ""),
                    safe(det.libro.descripcion_completa),
                    safe(det.libro.empresa), safe(det.libro.nivel), safe(det.libro.grado), safe(det.libro.area),
                    safe(det.cantidad_adoptada),
                    safe(det.mes_lectura or ""),
                ])

        for ws in [ws1, ws2]:
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col)].width = 22

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{ExportFilename("Reporte_General_BookExpress")}"'
        wb.save(response)
        return response
