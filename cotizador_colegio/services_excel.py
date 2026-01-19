from django.http import HttpResponse
from django.views import View
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from decimal import Decimal
import os
from datetime import datetime
from django.conf import settings

from .models import DetalleCotizacion, Adopcion


# ============================================================
# 🎨 PALETA CORPORATIVA
# ============================================================
COLOR_AZUL = "1F4E78"
COLOR_BLANCO = "FFFFFF"
COLOR_NEGRO = "000000"


def safe(value):
    if value is None:
        return ""
    # si es numero y es negativo, lo puedes dejar tal cual o clipear.
    # Si quieres mostrar negativos para ROI, comenta esto:
    if isinstance(value, (int, float, Decimal)) and value < 0:
        return float(value)  # ✅ no lo clamps a 0 (mejor para auditoría)
    return value


def ajustar_columnas(ws):
    for col in ws.columns:
        max_length = 0
        letter = None

        for cell in col:
            if hasattr(cell, "column_letter"):
                letter = cell.column_letter
                break

        if not letter:
            continue

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        ws.column_dimensions[letter].width = max_length + 2


def agregar_encabezado_keyfacil(ws, titulo):
    ws.insert_rows(1, amount=2)

    for row in range(1, 3):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(
                start_color=COLOR_AZUL,
                end_color=COLOR_AZUL,
                fill_type="solid"
            )

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28

    # Logo (ajusta si tu ruta real es diferente)
    logo_path = os.path.join(settings.BASE_DIR, "cotizador_colegio", "static", "img", "img_book_express.png")
    if os.path.exists(logo_path):
        img = XLImage(logo_path)
        img.width = 150
        img.height = 60
        ws.add_image(img, "A1")

    fecha = datetime.now().strftime("%d/%m/%Y")
    titulo_final = f"{titulo.upper()} – {fecha}"

    ws.merge_cells("D1:K2")
    c = ws["D1"]
    c.value = titulo_final
    c.font = Font(size=18, bold=True, color=COLOR_BLANCO)
    c.alignment = Alignment(horizontal="center", vertical="center")


def escribir_encabezados(ws, columnas):
    while ws.max_row < 3:
        ws.append([])

    ws.append(columnas)
    fila = ws.max_row

    for cell in ws[fila]:
        cell.fill = PatternFill(start_color=COLOR_AZUL, end_color=COLOR_AZUL, fill_type="solid")
        cell.font = Font(color=COLOR_BLANCO, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin", color=COLOR_NEGRO),
            right=Side(style="thin", color=COLOR_NEGRO),
            top=Side(style="thin", color=COLOR_NEGRO),
            bottom=Side(style="thin", color=COLOR_NEGRO),
        )


# ============================================================
# 📌 EXPORTAR COTIZACIONES
# ============================================================
class ExportCotizacionesExcelView(View):
    def get(self, request, *args, **kwargs):
        detalles = (
            DetalleCotizacion.objects
            .select_related("cotizacion", "cotizacion__asesor", "cotizacion__institucion", "producto", "producto__editorial")
            .all()
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Cotizaciones"

        agregar_encabezado_keyfacil(ws, "REPORTE DE COTIZACIONES")

        columnas = [
            "Editorial", "Nivel", "Grado", "Área", "Producto",
            "PVP 2026", "Tipo de Venta",
            "Precio BE", "% Desc Proveedor", "Precio Proveedor",
            "% Desc IE", "Dscto IE (S/)", "Precio IE",
            "Precio PPFF",
            "% Desc Consigna", "Comisión (S/)",
            "Utilidad IE", "ROI (S/)",
            "Asesor", "Institución", "Fecha",
        ]

        escribir_encabezados(ws, columnas)

        for d in detalles:
            cot = d.cotizacion
            prod = d.producto

            # texto tipo venta
            tv = (getattr(d, "tipo_venta", "") or "").upper()
            tv_map = {
                "PV": "Punto de Venta",
                "PUNTO_DE_VENTA": "Punto de Venta",
                "FERIA": "Feria",
                "CONSIGNA": "Consignación",
            }
            tv_txt = tv_map.get(tv, tv)

            precio_be = getattr(d, "precio_be", None)
            descuento_ie = getattr(d, "descuento_ie", None)
            descuento_ie_monto = None
            precio_ie = getattr(d, "precio_ie", None)

            if precio_be is not None and descuento_ie is not None:
                try:
                    descuento_ie_monto = (Decimal(str(precio_be)) * Decimal(str(descuento_ie)) / Decimal("100")).quantize(Decimal("0.01"))
                except Exception:
                    descuento_ie_monto = ""

            ws.append([
                safe(getattr(prod.editorial, "nombre", "")) if getattr(prod, "editorial", None) else "",
                safe(getattr(prod, "nivel", "")),
                safe(getattr(prod, "grado", "")),
                safe(getattr(prod, "area", "")),
                safe(getattr(prod, "nombre", "")),
                safe(getattr(prod, "pvp_2026", "")),
                safe(tv_txt),

                safe(precio_be),
                safe(getattr(d, "desc_proveedor", "")),
                safe(getattr(d, "precio_proveedor", "")),

                safe(descuento_ie),
                safe(descuento_ie_monto),
                safe(precio_ie),

                safe(getattr(d, "precio_ppff", "")),
                safe(getattr(d, "desc_consigna", "")),
                safe(getattr(d, "comision", "")),

                safe(getattr(d, "utilidad_ie", "")),
                safe(getattr(d, "roi_ie", "")),

                safe(getattr(cot.asesor, "nombre", "")) if getattr(cot, "asesor", None) else "",
                safe(getattr(cot.institucion, "nombre", "")) if getattr(cot, "institucion", None) else "",
                safe(cot.fecha.strftime("%d/%m/%Y") if getattr(cot, "fecha", None) else ""),
            ])

        ajustar_columnas(ws)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="Cotizaciones_{timezone.now().date()}.xlsx"'
        wb.save(response)
        return response


# ============================================================
# 📌 EXPORTAR ADOPCIONES
# ============================================================
class ExportAdopcionesExcelView(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        ws.title = "Adopciones"

        agregar_encabezado_keyfacil(ws, "REPORTE DE ADOPCIONES")

        columnas = [
            "N° Cotización", "Institución", "Asesor",
            "Editorial", "Producto", "Nivel", "Grado", "Área",
            "Cantidad", "Mes Lectura",
        ]
        escribir_encabezados(ws, columnas)

        adopciones = (
            Adopcion.objects
            .select_related("cotizacion", "cotizacion__institucion", "cotizacion__asesor")
            .prefetch_related("detalles__producto", "detalles__producto__editorial")
        )

        for adop in adopciones:
            for det in adop.detalles.all():
                prod = det.producto
                ws.append([
                    safe(getattr(adop.cotizacion, "numero_cotizacion", "")),
                    safe(getattr(adop.cotizacion.institucion, "nombre", "")) if getattr(adop.cotizacion, "institucion", None) else "",
                    safe(getattr(adop.cotizacion.asesor, "nombre", "")) if getattr(adop.cotizacion, "asesor", None) else "",

                    safe(getattr(prod.editorial, "nombre", "")) if getattr(prod, "editorial", None) else "",
                    safe(getattr(prod, "nombre", "")),
                    safe(getattr(prod, "nivel", "")),
                    safe(getattr(prod, "grado", "")),
                    safe(getattr(prod, "area", "")),

                    safe(getattr(det, "cantidad_adoptada", "")),
                    safe(getattr(det, "mes_lectura", "")),
                ])

        ajustar_columnas(ws)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="Adopciones_{timezone.now().date()}.xlsx"'
        wb.save(response)
        return response


# ============================================================
# 📌 EXPORTACIÓN GENERAL (2 HOJAS)
# ============================================================
class ExportGeneralExcelView(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()

        # Hoja 1: Cotizaciones
        ws1 = wb.active
        ws1.title = "Cotizaciones"
        agregar_encabezado_keyfacil(ws1, "REPORTE GENERAL – COTIZACIONES")

        columnas_cot = [
            "Editorial", "Nivel", "Grado", "Área", "Producto",
            "PVP 2026", "Tipo de Venta",
            "Precio BE", "% Desc Proveedor", "Precio Proveedor",
            "% Desc IE", "Dscto IE (S/)", "Precio IE",
            "Precio PPFF",
            "% Desc Consigna", "Comisión (S/)",
            "Utilidad IE", "ROI (S/)",
            "Asesor", "Institución", "Fecha",
        ]
        escribir_encabezados(ws1, columnas_cot)

        detalles = (
            DetalleCotizacion.objects
            .select_related("cotizacion", "cotizacion__asesor", "cotizacion__institucion", "producto", "producto__editorial")
            .all()
        )

        for d in detalles:
            cot = d.cotizacion
            prod = d.producto

            tv = (getattr(d, "tipo_venta", "") or "").upper()
            tv_map = {
                "PV": "Punto de Venta",
                "PUNTO_DE_VENTA": "Punto de Venta",
                "FERIA": "Feria",
                "CONSIGNA": "Consignación",
            }
            tv_txt = tv_map.get(tv, tv)

            precio_be = getattr(d, "precio_be", None)
            descuento_ie = getattr(d, "descuento_ie", None)
            descuento_ie_monto = None
            if precio_be is not None and descuento_ie is not None:
                try:
                    descuento_ie_monto = (Decimal(str(precio_be)) * Decimal(str(descuento_ie)) / Decimal("100")).quantize(Decimal("0.01"))
                except Exception:
                    descuento_ie_monto = ""

            ws1.append([
                safe(getattr(prod.editorial, "nombre", "")) if getattr(prod, "editorial", None) else "",
                safe(getattr(prod, "nivel", "")),
                safe(getattr(prod, "grado", "")),
                safe(getattr(prod, "area", "")),
                safe(getattr(prod, "nombre", "")),
                safe(getattr(prod, "pvp_2026", "")),
                safe(tv_txt),

                safe(getattr(d, "precio_be", "")),
                safe(getattr(d, "desc_proveedor", "")),
                safe(getattr(d, "precio_proveedor", "")),

                safe(getattr(d, "descuento_ie", "")),
                safe(descuento_ie_monto),
                safe(getattr(d, "precio_ie", "")),

                safe(getattr(d, "precio_ppff", "")),
                safe(getattr(d, "desc_consigna", "")),
                safe(getattr(d, "comision", "")),

                safe(getattr(d, "utilidad_ie", "")),
                safe(getattr(d, "roi_ie", "")),

                safe(getattr(cot.asesor, "nombre", "")) if getattr(cot, "asesor", None) else "",
                safe(getattr(cot.institucion, "nombre", "")) if getattr(cot, "institucion", None) else "",
                safe(cot.fecha.strftime("%d/%m/%Y") if getattr(cot, "fecha", None) else ""),
            ])

        ajustar_columnas(ws1)

        # Hoja 2: Adopciones
        ws2 = wb.create_sheet("Adopciones")
        agregar_encabezado_keyfacil(ws2, "REPORTE GENERAL – ADOPCIONES")

        columnas_adop = [
            "N° Cotización", "Institución", "Asesor",
            "Editorial", "Producto", "Nivel", "Grado", "Área",
            "Cantidad", "Mes Lectura",
        ]
        escribir_encabezados(ws2, columnas_adop)

        adopciones = (
            Adopcion.objects
            .select_related("cotizacion", "cotizacion__institucion", "cotizacion__asesor")
            .prefetch_related("detalles__producto", "detalles__producto__editorial")
        )

        for adop in adopciones:
            for det in adop.detalles.all():
                prod = det.producto
                ws2.append([
                    safe(getattr(adop.cotizacion, "numero_cotizacion", "")),
                    safe(getattr(adop.cotizacion.institucion, "nombre", "")) if getattr(adop.cotizacion, "institucion", None) else "",
                    safe(getattr(adop.cotizacion.asesor, "nombre", "")) if getattr(adop.cotizacion, "asesor", None) else "",

                    safe(getattr(prod.editorial, "nombre", "")) if getattr(prod, "editorial", None) else "",
                    safe(getattr(prod, "nombre", "")),
                    safe(getattr(prod, "nivel", "")),
                    safe(getattr(prod, "grado", "")),
                    safe(getattr(prod, "area", "")),

                    safe(getattr(det, "cantidad_adoptada", "")),
                    safe(getattr(det, "mes_lectura", "")),
                ])

        ajustar_columnas(ws2)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="ReporteGeneral_{timezone.now().date()}.xlsx"'
        wb.save(response)
        return response
